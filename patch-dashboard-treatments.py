#!/usr/bin/env python3
"""
patch-dashboard-treatments.py — ONE-OFF fix for an EXISTING her-expenses.xlsx.

The Dashboard's "Income by Treatment" table needs the treatment names in column L
(rows 4..21) for its per-treatment totals to populate. New workbooks get these from
treatments.json automatically (via create-workbook.py). This script back-fills the
names into a workbook that was built BEFORE that change — without rebuilding it, so
all her data and charts are preserved.

Safe to re-run: it only writes the name cells, nothing else. It makes a timestamped
backup first and never deletes anything.

Run:
    python patch-dashboard-treatments.py
"""
import os, sys, json, shutil
from datetime import datetime
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.environ.get("EXCEL_PATH", os.path.join(HERE, "her-expenses.xlsx"))
if not os.path.isabs(EXCEL_PATH):
    EXCEL_PATH = os.path.join(HERE, EXCEL_PATH)
TREATMENTS_PATH = os.environ.get("TREATMENTS_PATH", os.path.join(HERE, "treatments.json"))
if not os.path.isabs(TREATMENTS_PATH):
    TREATMENTS_PATH = os.path.join(HERE, TREATMENTS_PATH)

TREATMENT_SLOTS = 18
FIRST_ROW = 4          # Dashboard col L, first treatment name row
NAME_COL = 12          # column L


def load_names():
    with open(TREATMENTS_PATH, encoding="utf-8") as f:
        data = json.load(f)
    names, seen = [], set()
    for t in data:
        n = str(t.get("name", "")).strip()
        if n and n not in seen:
            seen.add(n); names.append(n)
    return names[:TREATMENT_SLOTS]


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"[!] {EXCEL_PATH} not found — nothing to patch.", file=sys.stderr)
        sys.exit(1)
    names = load_names()
    if not names:
        print("[!] No treatment names in treatments.json — nothing to write.", file=sys.stderr)
        sys.exit(1)

    # Backup first (never overwrite the only copy).
    stamp = datetime.now().strftime("%y%m%d-%H%M%S")
    backup = EXCEL_PATH.replace(".xlsx", f".backup-{stamp}.xlsx")
    shutil.copy2(EXCEL_PATH, backup)
    print(f"  Backup written: {os.path.basename(backup)}")

    wb = load_workbook(EXCEL_PATH)  # preserves charts + formulas
    if "Dashboard" not in wb.sheetnames:
        print("[!] No Dashboard sheet found.", file=sys.stderr)
        sys.exit(1)
    ws = wb["Dashboard"]

    for i in range(TREATMENT_SLOTS):
        r = FIRST_ROW + i
        ws.cell(row=r, column=NAME_COL, value=(names[i] if i < len(names) else ""))

    wb.save(EXCEL_PATH)
    print(f"  Wrote {len(names)} treatment name(s) into the Dashboard: {', '.join(names)}")
    print("  Open the workbook and let Excel recalculate — the Income-by-Treatment")
    print("  breakdown will now populate.")


if __name__ == "__main__":
    main()
