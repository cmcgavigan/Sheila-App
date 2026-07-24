"""v1 -> v2 data migration (audit A-01 / repair plan P0-1).

Imports the original app's data into the v2 SQLite database:
  - her-expenses.xlsx  Out/In sheets -> receipts (serials, dates, EUR values,
    statuses, notes, locations and photo links preserved)
  - her-photos\\*      -> data\\photos (SHA-256 verified copies)
  - treatments.json    -> treatments
  - places.json        -> places        (only if the file exists)
  - open-trip.json     -> open_trip     (only if the file exists)

Safety model:
  - The SOURCE is only ever read (workbook opened read_only). Never written.
  - DRY-RUN is the default and touches nothing; --apply is required to write.
  - The source directory is always explicit (--source). Nothing is assumed.
  - Apply refuses a target that already holds data (receipts/trips/treatments/
    open trip/photos) unless the recorded migration fingerprint matches the
    source, in which case the run is a no-op (idempotent re-run).
  - Apply is transactional: all rows go into one SQLite transaction; photos are
    staged, hash-verified, promoted into place, re-verified, and only then is
    the transaction committed. Any failure rolls back the rows and removes the
    promoted/staged files.
  - A zip backup of every source file is written into the target's backups
    directory before anything is imported, and verified against the hashes
    taken when the source was READ. If the source changes in between, the
    bad archive is deleted and the migration aborts — the imported rows and
    the backup can never represent two different source states.

Usage:
  python -m app.migrate --source "C:\\path\\to\\Sheilas app"            (dry run)
  python -m app.migrate --source "C:\\path\\to\\Sheilas app" --apply
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import os
import re
import shutil
import sqlite3
import sys
import tempfile
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

MIGRATION_VERSION = 1

WORKBOOK_NAME = "her-expenses.xlsx"
PHOTOS_DIRNAME = "her-photos"
TREATMENTS_NAME = "treatments.json"
PLACES_NAME = "places.json"
OPEN_TRIP_NAME = "open-trip.json"

# Column layouts of the v1 workbook (must match create-workbook.py exactly).
OUT_HEADERS = ["#", "Date", "Merchant", "Article", "Total", "Currency", "EUR Value",
               "Receipt Code", "Location", "Note", "Business/Personal", "Status",
               "Captured At", "Saved At"]
IN_HEADERS = ["#", "Date", "Client Name", "Treatment", "Total", "Currency", "EUR Value",
              "Receipt Code", "Location", "Note", "Status", "Captured At", "Saved At"]

_COORD_RE = re.compile(r"^-?\d+(?:\.\d+)?,\s*-?\d+(?:\.\d+)?$")


class MigrationError(Exception):
    """code 2 = source/validation problem, code 3 = target conflict/refusal."""

    def __init__(self, message: str, code: int = 2):
        super().__init__(message)
        self.code = code


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _norm_text(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_date(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _norm_float(v: Any, field: str, serial: int, sheet: str) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        raise MigrationError(
            f"{sheet} row #{serial}: {field} is not a number ({v!r})")


def _norm_eur(v: Any, serial: int, sheet: str) -> Optional[float]:
    if v in (None, ""):
        return None  # FX never resolved in v1 -> stays NULL (status CHECK FX)
    try:
        return float(v)
    except (TypeError, ValueError):
        raise MigrationError(
            f"{sheet} row #{serial}: EUR Value is not a number ({v!r})")


def _parse_serial(v: Any, sheet: str) -> int:
    try:
        n = int(str(v).strip())
    except (TypeError, ValueError):
        raise MigrationError(f"{sheet}: serial {v!r} is not an integer")
    if n < 1:
        raise MigrationError(f"{sheet}: serial {n} out of range")
    return n


# --------------------------------------------------------------------------
# source reading (pure — never writes anywhere)
# --------------------------------------------------------------------------

def _read_sheet(ws, expected_headers: list, kind: str, photo_names: set,
                warnings: list) -> list:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise MigrationError(f"{ws.title}: sheet is empty (no header row)")
    got = [c for c in header if c is not None]
    if list(got) != expected_headers:
        raise MigrationError(
            f"{ws.title}: header mismatch.\n  expected: {expected_headers}\n  found:    {list(got)}")

    prefix = "O" if kind == "out" else "I"
    out, seen = [], set()
    for i, raw in enumerate(rows_iter, start=2):
        vals = list(raw) + [None] * (len(expected_headers) - len(raw))
        if all(v in (None, "") for v in vals):
            continue
        if vals[0] in (None, ""):
            raise MigrationError(f"{ws.title} row {i}: data present but serial (#) empty")
        serial = _parse_serial(vals[0], ws.title)
        if serial in seen:
            raise MigrationError(f"{ws.title}: duplicate serial {serial}")
        seen.add(serial)

        location = _norm_text(vals[8])
        map_url = (f"https://www.google.com/maps?q={location.replace(' ', '')}"
                   if location and _COORD_RE.match(location) else "")
        photo_name = f"{prefix}{serial:03d}.jpg"
        has_photo = photo_name in photo_names
        if not has_photo:
            warnings.append(f"{ws.title} #{serial:03d}: no photo file ({photo_name}) — imported without photo link")

        row = {
            "kind": kind,
            "serial": serial,
            "date": _norm_date(vals[1]),
            "total": _norm_float(vals[4], "Total", serial, ws.title),
            "currency": _norm_text(vals[5]),
            "eur": _norm_eur(vals[6], serial, ws.title),
            "fx_rate": None,
            "fx_date": None,
            "receipt_code": _norm_text(vals[7]),
            "location": location,
            "map_url": map_url,
            "note": _norm_text(vals[9]),
            "photo": photo_name if has_photo else "",
            "generated_card": 0,
        }
        if kind == "out":
            row["merchant"] = _norm_text(vals[2])
            row["article"] = _norm_text(vals[3])
            row["business_personal"] = _norm_text(vals[10]) or "Business"
            row["status"] = _norm_text(vals[11]) or "OK"
            row["captured_at"] = _norm_text(vals[12])
            row["saved_at"] = _norm_text(vals[13])
        else:
            row["client_name"] = _norm_text(vals[2])
            row["treatment"] = _norm_text(vals[3])
            row["status"] = _norm_text(vals[10]) or "OK"
            row["captured_at"] = _norm_text(vals[11])
            row["saved_at"] = _norm_text(vals[12])
        out.append(row)
    out.sort(key=lambda r: r["serial"])
    return out


def _read_json_bytes(raw: bytes, what: str):
    try:
        return json.loads(raw.decode("utf-8"))
    except (UnicodeDecodeError, ValueError) as e:
        raise MigrationError(f"{what} is unreadable: {e}")


def read_source(source: Path) -> dict:
    """Read + validate everything from the v1 folder. Raises MigrationError.

    Every file is read into memory ONCE, hashed from those exact bytes, and
    parsed from those exact bytes. The hashes (file_hashes / photos) are what
    the backup archive and staged photo copies are later verified against, so
    a source file changing after this point is detected, not silently mixed in.
    Returns a dict with rows, photos (name -> (path, sha256)), lists, per-file
    read-time hashes and a fingerprint identifying this exact source state."""
    source = Path(source).resolve()
    if not source.is_dir():
        raise MigrationError(f"source folder not found: {source}")
    workbook = source / WORKBOOK_NAME
    if not workbook.is_file():
        raise MigrationError(f"{WORKBOOK_NAME} not found in {source} — is this really the v1 app folder?")

    warnings: list = []
    file_hashes: dict = {}

    photos: dict = {}
    photos_dir = source / PHOTOS_DIRNAME
    if photos_dir.is_dir():
        for p in sorted(photos_dir.iterdir()):
            if p.is_file():
                photos[p.name] = (p, _sha256_file(p))
    else:
        warnings.append(f"{PHOTOS_DIRNAME} folder not found — no photos to migrate")

    try:
        wb_bytes = workbook.read_bytes()
    except OSError as e:
        raise MigrationError(f"{WORKBOOK_NAME} is unreadable: {e}")
    file_hashes[WORKBOOK_NAME] = hashlib.sha256(wb_bytes).hexdigest()

    from openpyxl import load_workbook  # heavy import, keep it local
    wb = load_workbook(io.BytesIO(wb_bytes), read_only=True, data_only=True)
    try:
        for sheet in ("Out", "In"):
            if sheet not in wb.sheetnames:
                raise MigrationError(f"workbook has no '{sheet}' sheet: {wb.sheetnames}")
        out_rows = _read_sheet(wb["Out"], OUT_HEADERS, "out", set(photos), warnings)
        in_rows = _read_sheet(wb["In"], IN_HEADERS, "in", set(photos), warnings)
    finally:
        wb.close()

    def _load_json_file(name: str, what: str):
        p = source / name
        if not p.is_file():
            return None, False
        try:
            raw = p.read_bytes()
        except OSError as e:
            raise MigrationError(f"{what} ({name}) is unreadable: {e}")
        file_hashes[name] = hashlib.sha256(raw).hexdigest()
        return _read_json_bytes(raw, f"{what} ({name})"), True

    treatments: list = []
    data, present = _load_json_file(TREATMENTS_NAME, "treatments")
    if present:
        if not isinstance(data, list):
            raise MigrationError(f"{TREATMENTS_NAME} must contain a JSON array")
        for t in data:
            name = _norm_text((t or {}).get("name"))
            if name:
                try:
                    price = float((t or {}).get("price") or 0)
                except (TypeError, ValueError):
                    price = 0.0
                treatments.append({"name": name, "price": price})
    else:
        warnings.append(f"{TREATMENTS_NAME} not found — no treatments to migrate")

    places = None
    data, present = _load_json_file(PLACES_NAME, "places")
    if present:
        if not isinstance(data, list):
            raise MigrationError(f"{PLACES_NAME} must contain a JSON array")
        places = []
        for p in data:
            name = _norm_text((p or {}).get("name"))
            if name:
                places.append({"name": name, "address": _norm_text((p or {}).get("address"))})

    open_trip = None
    data, present = _load_json_file(OPEN_TRIP_NAME, "open trip")
    if present:
        if isinstance(data, dict) and data:
            open_trip = data
        elif data:
            raise MigrationError(f"{OPEN_TRIP_NAME} must contain a JSON object")

    referenced = {r["photo"] for r in out_rows + in_rows if r["photo"]}
    for name in photos:
        if name not in referenced:
            warnings.append(f"photo {name} is not referenced by any row — copied anyway")

    # Built purely from the read-time hashes — never re-hashed from disk.
    fingerprint = hashlib.sha256(json.dumps({
        "version": MIGRATION_VERSION,
        "workbook": file_hashes[WORKBOOK_NAME],
        "treatments": file_hashes.get(TREATMENTS_NAME, ""),
        "places": file_hashes.get(PLACES_NAME, ""),
        "open_trip": file_hashes.get(OPEN_TRIP_NAME, ""),
        "photos": sorted((n, sha) for n, (_p, sha) in photos.items()),
    }, sort_keys=True).encode()).hexdigest()

    def _eur_sum(rows):
        return round(sum(r["eur"] for r in rows if r["eur"] is not None), 2)

    return {
        "source": source,
        "out_rows": out_rows,
        "in_rows": in_rows,
        "treatments": treatments,
        "places": places,
        "open_trip": open_trip,
        "photos": photos,
        "file_hashes": file_hashes,
        "fingerprint": fingerprint,
        "warnings": warnings,
        "counts": {
            "out": len(out_rows),
            "in": len(in_rows),
            "photos": len(photos),
            "treatments": len(treatments),
            "places": len(places) if places is not None else 0,
            "open_trip": 1 if open_trip else 0,
            "eur_out": _eur_sum(out_rows),
            "eur_in": _eur_sum(in_rows),
        },
    }


# --------------------------------------------------------------------------
# apply
# --------------------------------------------------------------------------

def _connect_target(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path, timeout=15)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _meta(conn: sqlite3.Connection, key: str) -> str:
    row = conn.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else ""


def _refuse_if_target_conflicts(conn: sqlite3.Connection, src: dict,
                                photos_dir: Path) -> None:
    receipts = conn.execute("SELECT COUNT(*) c FROM receipts").fetchone()["c"]
    trips = conn.execute("SELECT COUNT(*) c FROM trips").fetchone()["c"]
    open_trip = conn.execute("SELECT COUNT(*) c FROM open_trip").fetchone()["c"]
    if receipts or trips or open_trip:
        raise MigrationError(
            f"target database already contains data (receipts={receipts}, trips={trips}, "
            f"open_trip={open_trip}) and its migration fingerprint does not match this "
            "source. Refusing — migrate into an empty v2 data directory.", code=3)
    if src["treatments"]:
        n = conn.execute("SELECT COUNT(*) c FROM treatments").fetchone()["c"]
        if n:
            raise MigrationError(
                f"target already has {n} treatments. Refusing to overwrite them.", code=3)
    if src["places"] is not None:
        rows = conn.execute("SELECT name FROM places ORDER BY pos, id").fetchall()
        names = [r["name"] for r in rows]
        if names and names != ["Zuhause"]:
            raise MigrationError(
                f"target already has places {names}. Refusing to overwrite them.", code=3)
    if photos_dir.is_dir():
        existing = [p.name for p in photos_dir.iterdir() if p.is_file()]
        if existing:
            raise MigrationError(
                f"target photos directory already contains {len(existing)} file(s). "
                "Refusing — migrate into an empty v2 data directory.", code=3)


def _write_backup(src: dict, backups_dir: Path) -> Path:
    """Zip every source file into the target's backups dir, then verify the
    archive byte-for-byte against the hashes taken at READ time (read_source).

    This is the TOCTOU guard: if any source file changed between being read
    and being archived, verification fails, the unusable archive is deleted
    and the migration aborts before a single row is imported — the imported
    data and the backup can therefore never represent different source states.
    """
    backups_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    zpath = backups_dir / f"pre-migration-{stamp}-{src['fingerprint'][:8]}.zip"

    members: list = []  # (arcname, source path, read-time sha256)
    source = src["source"]
    for name in (WORKBOOK_NAME, TREATMENTS_NAME, PLACES_NAME, OPEN_TRIP_NAME):
        sha = src["file_hashes"].get(name)
        if sha:
            members.append((name, source / name, sha))
    for name, (p, sha) in sorted(src["photos"].items()):
        members.append((f"{PHOTOS_DIRNAME}/{name}", p, sha))

    try:
        with zipfile.ZipFile(zpath, "w", zipfile.ZIP_DEFLATED) as z:
            for arcname, p, _sha in members:
                try:
                    z.write(p, arcname)
                except OSError as e:
                    raise MigrationError(
                        f"{arcname} vanished or became unreadable while writing the "
                        f"backup ({e}) — is the v1 app still running? Stop it and re-run.")
        with zipfile.ZipFile(zpath) as z:
            for arcname, _p, sha in members:
                if hashlib.sha256(z.read(arcname)).hexdigest() != sha:
                    raise MigrationError(
                        f"{arcname} changed on disk between reading and backup — the "
                        "source is being modified concurrently. Stop the v1 app and "
                        "re-run the migration.")
    except BaseException:
        zpath.unlink(missing_ok=True)  # a mismatched backup must not survive
        raise
    return zpath


RECEIPT_COLS = ["kind", "serial", "date", "merchant", "article", "business_personal",
                "client_name", "treatment", "total", "currency", "eur", "fx_rate",
                "fx_date", "receipt_code", "note", "location", "map_url", "status",
                "photo", "generated_card", "captured_at", "saved_at"]


def apply_migration(src: dict, data_dir: Path) -> dict:
    from . import db as appdb  # imported lazily; pulls in app.config

    data_dir = Path(data_dir).resolve()
    source = src["source"]
    if source == data_dir or source in data_dir.parents or data_dir in source.parents:
        raise MigrationError("source and target directories overlap — refusing", code=3)

    photos_dir = data_dir / "photos"
    data_dir.mkdir(parents=True, exist_ok=True)

    conn = _connect_target(data_dir / "sheila.db")
    try:
        appdb.init_schema(conn)
        conn.commit()

        existing_fp = _meta(conn, "migration.source_fingerprint")
        if existing_fp == src["fingerprint"]:
            return {"applied": False, "already_migrated": True,
                    "fingerprint": src["fingerprint"], "counts": src["counts"],
                    "warnings": src["warnings"]}
        if existing_fp:
            raise MigrationError(
                "target was already migrated from a DIFFERENT source state "
                f"(recorded {existing_fp[:12]}…, this source {src['fingerprint'][:12]}…). "
                "Refusing. Use a fresh v2 data directory, or re-run against the "
                "identical source.", code=3)
        _refuse_if_target_conflicts(conn, src, photos_dir)

        backup = _write_backup(src, data_dir / "backups")

        # mkdtemp gives this run its own private staging dir; the cleanup in
        # the finally below removes only that exact directory — a pre-existing
        # path can never be deleted.
        staging_dir = Path(tempfile.mkdtemp(prefix=".migration-staging-", dir=data_dir))
        photos_dir.mkdir(parents=True, exist_ok=True)
        staged: dict = {}
        moved: list = []
        try:
            # Stage photo copies and verify them against the read-time hashes.
            for name, (p, sha) in src["photos"].items():
                sp = staging_dir / name
                try:
                    shutil.copyfile(p, sp)
                except OSError as e:
                    raise MigrationError(f"could not stage photo {name}: {e}")
                if _sha256_file(sp) != sha:
                    raise MigrationError(
                        f"photo {name} changed on disk between reading and staging — "
                        "the source is being modified concurrently. Stop the v1 app "
                        "and re-run the migration.")
                staged[name] = (sp, sha)

            # All rows in ONE transaction; commit happens only after the
            # photos are promoted and re-verified.
            marks = ",".join("?" for _ in RECEIPT_COLS)
            for row in src["out_rows"] + src["in_rows"]:
                conn.execute(
                    f"INSERT INTO receipts ({', '.join(RECEIPT_COLS)}) VALUES ({marks})",
                    [row.get(c) for c in RECEIPT_COLS])
            for i, t in enumerate(src["treatments"]):
                conn.execute("INSERT INTO treatments(name, price, pos) VALUES (?,?,?)",
                             (t["name"], t["price"], i))
            if src["places"] is not None:
                conn.execute("DELETE FROM places")
                for i, p in enumerate(src["places"]):
                    conn.execute("INSERT INTO places(name, address, pos) VALUES (?,?,?)",
                                 (p["name"], p["address"], i))
            if src["open_trip"]:
                conn.execute(
                    "INSERT INTO open_trip(id, payload) VALUES (1, ?)",
                    (json.dumps(src["open_trip"], ensure_ascii=False),))
            conn.execute(
                "INSERT OR REPLACE INTO meta(key, value) VALUES "
                "('migration.version', ?), ('migration.source', ?), "
                "('migration.source_fingerprint', ?), ('migration.counts', ?), "
                "('migration.timestamp', ?), ('migration.backup', ?)",
                (str(MIGRATION_VERSION), str(source), src["fingerprint"],
                 json.dumps(src["counts"]), datetime.now().isoformat(timespec="seconds"),
                 backup.name))

            # Reconcile inside the open transaction, before anything is final.
            for kind, rows in (("out", src["out_rows"]), ("in", src["in_rows"])):
                got = conn.execute(
                    "SELECT COUNT(*) c, ROUND(COALESCE(SUM(eur), 0), 2) s "
                    "FROM receipts WHERE kind = ?", (kind,)).fetchone()
                want_n = len(rows)
                want_s = round(sum(r["eur"] for r in rows if r["eur"] is not None), 2)
                if got["c"] != want_n or abs(got["s"] - want_s) > 0.005:
                    raise MigrationError(
                        f"reconciliation failed for '{kind}': rows {got['c']}/{want_n}, "
                        f"EUR {got['s']}/{want_s}")

            # Promote photos into place, re-verify each in its final location.
            for name, (sp, sha) in sorted(staged.items()):
                dest = photos_dir / name
                os.replace(sp, dest)
                moved.append(dest)
                if _sha256_file(dest) != sha:
                    raise MigrationError(f"promoted photo {name} failed hash verification")

            conn.commit()
        except BaseException:
            conn.rollback()
            for dest in moved:
                try:
                    dest.unlink()
                except OSError:
                    pass
            raise
        finally:
            shutil.rmtree(staging_dir, ignore_errors=True)

        return {"applied": True, "already_migrated": False,
                "fingerprint": src["fingerprint"], "counts": src["counts"],
                "warnings": src["warnings"], "backup": str(backup)}
    finally:
        conn.close()


# --------------------------------------------------------------------------
# entry points
# --------------------------------------------------------------------------

def run(source, data_dir=None, apply_changes: bool = False) -> dict:
    """Programmatic entry point. Dry-run unless apply_changes=True."""
    src = read_source(Path(source))
    if not apply_changes:
        return {"applied": False, "already_migrated": False, "dry_run": True,
                "fingerprint": src["fingerprint"], "counts": src["counts"],
                "warnings": src["warnings"]}
    if data_dir is None:
        from . import config  # lazy: creates the app's data dirs on import
        data_dir = config.DATA_DIR
    else:
        # app.config creates its data directories when first imported (via
        # app.db inside apply_migration). Point that side effect at the
        # explicit target so a CLI run never creates the app's default data
        # dir. Inert if app.config is already loaded (e.g. under pytest).
        os.environ.setdefault("DATA_DIR", str(data_dir))
    return apply_migration(src, Path(data_dir))


def _print_report(report: dict, source: str) -> None:
    c = report["counts"]
    mode = ("ALREADY MIGRATED — nothing to do" if report.get("already_migrated")
            else "APPLIED" if report.get("applied") else "DRY RUN — nothing was written")
    print(f"\n  v1 -> v2 migration  [{mode}]")
    print(f"  source:       {source}")
    print(f"  fingerprint:  {report['fingerprint'][:16]}…")
    print(f"  expenses:     {c['out']} rows  (EUR sum {c['eur_out']:.2f})")
    print(f"  income:       {c['in']} rows  (EUR sum {c['eur_in']:.2f})")
    print(f"  photos:       {c['photos']}")
    print(f"  treatments:   {c['treatments']}")
    print(f"  places:       {c['places'] if c['places'] else '(none — keeping v2 default)'}")
    print(f"  open trip:    {'yes' if c['open_trip'] else 'no'}")
    if report.get("backup"):
        print(f"  backup:       {report['backup']}")
    for w in report["warnings"]:
        print(f"  [!] {w}")
    if report.get("dry_run"):
        print("\n  Nothing was changed. Re-run with --apply to import.")


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(
        prog="python -m app.migrate",
        description="Migrate v1 (workbook) data into the v2 database. Dry-run by default.")
    ap.add_argument("--source", required=True,
                    help="path to the v1 app folder (the one containing her-expenses.xlsx)")
    ap.add_argument("--data-dir", default=None,
                    help="target v2 data directory (default: the app's configured data dir)")
    ap.add_argument("--apply", action="store_true",
                    help="actually write. Without this flag nothing is modified.")
    args = ap.parse_args(argv)

    try:
        report = run(args.source, data_dir=args.data_dir, apply_changes=args.apply)
    except MigrationError as e:
        print(f"\n  [!!] {e}", file=sys.stderr)
        return e.code
    _print_report(report, args.source)
    return 0


if __name__ == "__main__":
    sys.exit(main())
