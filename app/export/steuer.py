"""Steuerberater workbook generator — the ONLY place Excel is written.

Generates a fresh, complete workbook from SQLite on every call:
  Info      - business details, §19 note, tax rates, deductibility table
  Reisen    - one row per business trip, live formulas referencing Info rates
  Fahrten   - one row per trip leg
  Ausgaben  - expenses with auto-deductibility (VLOOKUP into the Info table)
  Einnahmen - income
  Dashboard - native Excel charts over monthly aggregates

Layout follows the version agreed on 2026-07-24 (bilingual DE(EN) headers,
Kleinunternehmerin §19 UStG so no VAT columns). When Herr Schwarz replies with
format wishes, THIS file is the only thing that changes — the data never moves.
"""
from __future__ import annotations

import os
import tempfile
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .. import config, db

ACCENT = "B6356B"      # app pink
HEAD_FILL = PatternFill("solid", fgColor="F3E1EB")
HEAD_FONT = Font(bold=True, color="7A1F4A")
TITLE_FONT = Font(bold=True, size=14, color=ACCENT)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(bottom=THIN)
EUR_FMT = '#,##0.00 "€"'
PCT_FMT = "0%"

# Fixed anchors on the Info sheet, referenced by formulas elsewhere.
RATE_KM = "Info!$B$6"
RATE_V8 = "Info!$B$7"
RATE_V24 = "Info!$B$8"
CLASS_FIRST_ROW = 13  # deductibility table data starts here (col A/B/C)


def safe_excel_text(value):
    """Return user/OCR text as a literal Excel value, never a formula."""
    if not isinstance(value, str):
        return value
    dangerous = ("=", "+", "-", "@", "\t", "\r", "\n")
    return "'" + value if value.startswith(dangerous) else value


def excel_date(value):
    """Convert stored ISO dates/datetimes into real Excel date values."""
    if not value:
        return value
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed if "T" in str(value) else parsed.date()
    except (TypeError, ValueError):
        return safe_excel_text(str(value))


def _head(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDER
        if widths and i <= len(widths) and widths[i - 1]:
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _hours(start: str, end: str) -> float:
    try:
        s = datetime.fromisoformat(start)
        e = datetime.fromisoformat(end)
        return max(0.0, round((e - s).total_seconds() / 3600, 1))
    except (ValueError, TypeError):
        return 0.0


def _month_key(d: str) -> str:
    return str(d or "")[:7] if d else ""


def generate_workbook(out_path: Path = None) -> Path:
    with db.get_db() as conn:
        rates = db.tax_rates(conn)
        outs = conn.execute(
            "SELECT * FROM receipts WHERE kind='out' ORDER BY date, serial").fetchall()
        ins = conn.execute(
            "SELECT * FROM receipts WHERE kind='in' ORDER BY date, serial").fetchall()
        trips = conn.execute("SELECT * FROM trips ORDER BY start, id").fetchall()
        legs_by_trip = defaultdict(list)
        for leg in conn.execute("SELECT * FROM legs ORDER BY trip_id, seq").fetchall():
            legs_by_trip[leg["trip_id"]].append(leg)

    wb = Workbook()

    # ================= Info =================
    ws = wb.active
    ws.title = "Info"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 58
    ws["A1"] = "Sheila McGavigan — Permanent Makeup"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Kleinunternehmerin gem. §19 UStG — kein Umsatzsteuerausweis (no VAT shown)"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = f"Export erstellt (generated): {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A3"].font = Font(color="666666")

    ws["A5"] = "Pauschalen (flat rates)"
    ws["A5"].font = Font(bold=True)
    ws["A6"] = "Kilometerpauschale €/km (mileage rate)"
    ws["B6"] = rates["rate_km"]
    ws["A7"] = "Verpflegung eintägig >8h (per-diem, single day)"
    ws["B7"] = rates["rate_meal_partial"]
    ws["A8"] = "Verpflegung mehrtägig je Übernachtung (per-diem, multi-day)"
    ws["B8"] = rates["rate_meal_full"]
    ws["A9"] = "Kürzung Frühstück (breakfast reduction)"
    ws["B9"] = rates["breakfast_reduction"]
    for r in range(6, 10):
        ws.cell(row=r, column=2).number_format = EUR_FMT

    ws["A11"] = "Abzugsfähigkeit nach Kategorie (deductibility by category)"
    ws["A11"].font = Font(bold=True)
    ws["A12"], ws["B12"], ws["C12"] = "Kategorie (category)", "%", "Behandlung (treatment)"
    for col in "ABC":
        ws[f"{col}12"].fill = HEAD_FILL
        ws[f"{col}12"].font = HEAD_FONT
    r = CLASS_FIRST_ROW
    for art in config.ARTICLES:
        pct, note = config.DEDUCTIBILITY.get(art, config.DEDUCTIBILITY_DEFAULT)
        ws.cell(row=r, column=1, value=art)
        c = ws.cell(row=r, column=2, value=pct / 100)
        c.number_format = PCT_FMT
        ws.cell(row=r, column=3, value=note)
        r += 1
    class_last = r - 1
    class_range = f"Info!$A${CLASS_FIRST_ROW}:$B${class_last}"

    # ================= Reisen =================
    ws = wb.create_sheet("Reisen")
    headers = ["Nr", "Zweck (purpose)", "Abfahrt (depart)", "Rückkehr (return)",
               "Route", "Nächte (nights)", "Std (hours)", "km",
               "Kilometer € (mileage)", "Verpflegung € (per-diem)",
               "abzgl. Mahlzeiten € (less meals)", "Übernachtung € (lodging)",
               "Nebenkosten € (incidentals)", "Summe € (total)", "Notiz (note)"]
    _head(ws, 1, headers,
          widths=[5, 22, 16, 16, 40, 8, 8, 9, 12, 13, 13, 13, 13, 12, 24])
    rr = 2
    for i, t in enumerate(trips, start=1):
        legs = legs_by_trip[t["id"]]
        route = " → ".join(
            [legs[0]["from_place"]] + [l["to_place"] for l in legs]) if legs else ""
        km_total = round(sum(l["km"] or 0 for l in legs), 1)
        ws.cell(row=rr, column=1, value=i)
        ws.cell(row=rr, column=2, value=safe_excel_text(t["purpose"]))
        ws.cell(row=rr, column=3, value=excel_date(t["start"]))
        ws.cell(row=rr, column=4, value=excel_date(t["end"]))
        ws.cell(row=rr, column=3).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=rr, column=4).number_format = "yyyy-mm-dd hh:mm"
        ws.cell(row=rr, column=5, value=safe_excel_text(route))
        ws.cell(row=rr, column=6, value=t["nights"])
        ws.cell(row=rr, column=7, value=_hours(t["start"], t["end"]))
        ws.cell(row=rr, column=8, value=km_total)
        ws.cell(row=rr, column=9, value=f"=H{rr}*{RATE_KM}")
        ws.cell(row=rr, column=10,
                value=f"=IF(F{rr}>0,{RATE_V24}*F{rr},IF(G{rr}>8,{RATE_V8},0))")
        ws.cell(row=rr, column=11, value=t["meal_reduction"] or 0)
        ws.cell(row=rr, column=12, value=t["overnight_cost"] or 0)
        ws.cell(row=rr, column=13, value=t["incidentals"] or 0)
        ws.cell(row=rr, column=14, value=f"=I{rr}+J{rr}-K{rr}+L{rr}+M{rr}")
        ws.cell(row=rr, column=15, value=safe_excel_text(t["note"]))
        for col in range(9, 15):
            ws.cell(row=rr, column=col).number_format = EUR_FMT
        rr += 1
    if rr > 2:
        ws.cell(row=rr + 1, column=8, value=f"=SUM(H2:H{rr - 1})").font = Font(bold=True)
        s = ws.cell(row=rr + 1, column=14, value=f"=SUM(N2:N{rr - 1})")
        s.font = Font(bold=True)
        s.number_format = EUR_FMT
        ws.cell(row=rr + 1, column=2, value="Summe (total)").font = Font(bold=True)

    # ================= Fahrten =================
    ws = wb.create_sheet("Fahrten")
    _head(ws, 1, ["Reise-Nr (trip)", "Datum (date)", "Von (from)", "Nach (to)", "km"],
          widths=[10, 12, 26, 26, 9])
    rr = 2
    for i, t in enumerate(trips, start=1):
        for leg in legs_by_trip[t["id"]]:
            ws.cell(row=rr, column=1, value=i)
            ws.cell(row=rr, column=2, value=excel_date(leg["date"]))
            ws.cell(row=rr, column=2).number_format = "yyyy-mm-dd"
            ws.cell(row=rr, column=3, value=safe_excel_text(leg["from_place"]))
            ws.cell(row=rr, column=4, value=safe_excel_text(leg["to_place"]))
            ws.cell(row=rr, column=5, value=leg["km"])
            rr += 1

    # ================= Ausgaben =================
    ws = wb.create_sheet("Ausgaben")
    headers = ["Nr", "Datum (date)", "Händler (merchant)", "Kategorie (category)",
               "Betrag (amount)", "Währung (curr.)", "EUR",
               "abziehbar % (deductible)", "abziehbar € (deductible)",
               "Geschäftlich/Privat", "Beleg-Nr (receipt)", "Status", "Notiz (note)"]
    _head(ws, 1, headers, widths=[6, 12, 26, 24, 11, 8, 11, 11, 12, 14, 14, 10, 24])
    rr = 2
    for o in outs:
        label = f"O{o['serial']}"
        c = ws.cell(row=rr, column=1, value=label)
        photo = config.PHOTOS_DIR / (o["photo"] or "")
        if o["photo"] and photo.exists():
            c.hyperlink = f"..\\photos\\{o['photo']}"
            c.font = Font(color="0563C1", underline="single")
        ws.cell(row=rr, column=2, value=excel_date(o["date"]))
        ws.cell(row=rr, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=rr, column=3, value=safe_excel_text(o["merchant"]))
        ws.cell(row=rr, column=4, value=safe_excel_text(o["article"]))
        ws.cell(row=rr, column=5, value=o["total"])
        ws.cell(row=rr, column=6, value=safe_excel_text(o["currency"]))
        ws.cell(row=rr, column=7, value=o["eur"])
        is_personal = (o["business_personal"] == "Personal")
        if is_personal:
            ws.cell(row=rr, column=8, value=0).number_format = PCT_FMT
        else:
            pc = ws.cell(row=rr, column=8,
                         value=f"=IFERROR(VLOOKUP(D{rr},{class_range},2,FALSE),1)")
            pc.number_format = PCT_FMT
        ws.cell(row=rr, column=9, value=f"=IF(G{rr}=\"\",0,G{rr})*H{rr}").number_format = EUR_FMT
        ws.cell(row=rr, column=10, value="Privat" if is_personal else "Geschäftlich")
        ws.cell(row=rr, column=11, value=safe_excel_text(o["receipt_code"]))
        ws.cell(row=rr, column=12, value=safe_excel_text(o["status"]))
        ws.cell(row=rr, column=13, value=safe_excel_text(o["note"]))
        ws.cell(row=rr, column=5).number_format = "#,##0.00"
        ws.cell(row=rr, column=7).number_format = EUR_FMT
        rr += 1
    out_last = rr - 1
    if outs:
        ws.cell(row=rr + 1, column=3, value="Summe abziehbar (total deductible)").font = Font(bold=True)
        s = ws.cell(row=rr + 1, column=9, value=f"=SUM(I2:I{out_last})")
        s.font = Font(bold=True)
        s.number_format = EUR_FMT

    # ================= Einnahmen =================
    ws = wb.create_sheet("Einnahmen")
    headers = ["Nr", "Datum (date)", "Kundin (client)", "Behandlung (treatment)",
               "Betrag (amount)", "Währung (curr.)", "EUR", "Beleg-Nr (receipt)",
               "Status", "Notiz (note)"]
    _head(ws, 1, headers, widths=[6, 12, 24, 26, 11, 8, 11, 14, 10, 24])
    rr = 2
    for i_ in ins:
        label = f"I{i_['serial']}"
        c = ws.cell(row=rr, column=1, value=label)
        photo = config.PHOTOS_DIR / (i_["photo"] or "")
        if i_["photo"] and photo.exists():
            c.hyperlink = f"..\\photos\\{i_['photo']}"
            c.font = Font(color="0563C1", underline="single")
        ws.cell(row=rr, column=2, value=excel_date(i_["date"]))
        ws.cell(row=rr, column=2).number_format = "yyyy-mm-dd"
        ws.cell(row=rr, column=3, value=safe_excel_text(i_["client_name"]))
        ws.cell(row=rr, column=4, value=safe_excel_text(i_["treatment"]))
        ws.cell(row=rr, column=5, value=i_["total"]).number_format = "#,##0.00"
        ws.cell(row=rr, column=6, value=safe_excel_text(i_["currency"]))
        ws.cell(row=rr, column=7, value=i_["eur"]).number_format = EUR_FMT
        ws.cell(row=rr, column=8, value=safe_excel_text(i_["receipt_code"]))
        ws.cell(row=rr, column=9, value=safe_excel_text(i_["status"]))
        ws.cell(row=rr, column=10, value=safe_excel_text(i_["note"]))
        rr += 1
    in_last = rr - 1
    if ins:
        ws.cell(row=rr + 1, column=4, value="Summe (total)").font = Font(bold=True)
        s = ws.cell(row=rr + 1, column=7, value=f"=SUM(G2:G{in_last})")
        s.font = Font(bold=True)
        s.number_format = EUR_FMT

    # ================= Dashboard =================
    _build_dashboard(wb, outs, ins, trips, legs_by_trip, rates)

    # ================= save =================
    if out_path is None:
        stamp = datetime.now().strftime("%y%m%d-%H%M%S")
        out_path = config.EXPORTS_DIR / f"{stamp}-{uuid.uuid4().hex[:10]}_sheila_steuer_export.xlsx"
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix=f".{out_path.stem}-", suffix=".tmp",
                                      dir=out_path.parent)
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        os.replace(temp_path, out_path)
    finally:
        temp_path.unlink(missing_ok=True)
    return out_path


def _build_dashboard(wb, outs, ins, trips, legs_by_trip, rates):
    ws = wb.create_sheet("Dashboard")
    ws["A1"] = "Dashboard"
    ws["A1"].font = TITLE_FONT
    ws.column_dimensions["A"].width = 12
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    # ---- monthly aggregates (computed in Python; export regenerates anyway) ----
    months = sorted({m for m in (
        [_month_key(o["date"]) for o in outs] +
        [_month_key(i["date"]) for i in ins] +
        [_month_key(t["start"]) for t in trips]) if m})
    inc = defaultdict(float)
    exp = defaultdict(float)
    deductible = defaultdict(float)
    for i_ in ins:
        if i_["eur"]:
            inc[_month_key(i_["date"])] += i_["eur"]
    for o in outs:
        if o["eur"] and o["business_personal"] != "Personal":
            exp[_month_key(o["date"])] += o["eur"]
            pct = config.DEDUCTIBILITY.get(
                o["article"], config.DEDUCTIBILITY_DEFAULT
            )[0] / 100
            deductible[_month_key(o["date"])] += o["eur"] * pct
    # trips count as expenses too (mileage + per-diem etc.)
    for t in trips:
        legs = legs_by_trip[t["id"]]
        km = sum(l["km"] or 0 for l in legs)
        per_diem = (rates["rate_meal_full"] * t["nights"] if t["nights"] > 0
                    else (rates["rate_meal_partial"] if _hours(t["start"], t["end"]) > 8 else 0))
        total = (km * rates["rate_km"] + per_diem - (t["meal_reduction"] or 0)
                 + (t["overnight_cost"] or 0) + (t["incidentals"] or 0))
        exp[_month_key(t["start"])] += total
        deductible[_month_key(t["start"])] += total

    ws["A3"], ws["B3"], ws["C3"], ws["D3"], ws["E3"] = (
        "Monat (month)", "Einnahmen €", "Ausgaben €", "Netto €",
        "Abziehbar € (deductible)")
    for col in "ABCDE":
        ws[f"{col}3"].fill = HEAD_FILL
        ws[f"{col}3"].font = HEAD_FONT
    r = 4
    for m in months:
        ws.cell(row=r, column=1, value=m)
        ws.cell(row=r, column=2, value=round(inc.get(m, 0), 2)).number_format = EUR_FMT
        ws.cell(row=r, column=3, value=round(exp.get(m, 0), 2)).number_format = EUR_FMT
        ws.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = EUR_FMT
        ws.cell(row=r, column=5, value=round(deductible.get(m, 0), 2)).number_format = EUR_FMT
        r += 1
    last = r - 1

    # cumulative net (true running total — v1's REVISIT item, done properly here)
    ws["F3"] = "Kumuliert € (cumulative net)"
    ws["F3"].fill = HEAD_FILL
    ws["F3"].font = HEAD_FONT
    for rr in range(4, last + 1):
        ws.cell(row=rr, column=6,
                value=f"=SUM($D$4:D{rr})").number_format = EUR_FMT

    # income by treatment
    by_treat = defaultdict(float)
    for i_ in ins:
        if i_["eur"]:
            by_treat[(i_["treatment"] or "—").strip() or "—"] += i_["eur"]
    ws["H3"], ws["I3"] = "Behandlung (treatment)", "Einnahmen €"
    ws["H3"].fill = HEAD_FILL
    ws["H3"].font = HEAD_FONT
    ws["I3"].fill = HEAD_FILL
    ws["I3"].font = HEAD_FONT
    ws.column_dimensions["H"].width = 26
    r = 4
    for name, total in sorted(by_treat.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=8, value=safe_excel_text(name))
        ws.cell(row=r, column=9, value=round(total, 2)).number_format = EUR_FMT
        r += 1
    treat_last = r - 1

    # expenses by category
    by_cat = defaultdict(float)
    for o in outs:
        if o["eur"] and o["business_personal"] != "Personal":
            by_cat[o["article"] or "—"] += o["eur"]
    ws["K3"], ws["L3"] = "Kategorie (category)", "Ausgaben €"
    ws["K3"].fill = HEAD_FILL
    ws["K3"].font = HEAD_FONT
    ws["L3"].fill = HEAD_FILL
    ws["L3"].font = HEAD_FONT
    ws.column_dimensions["K"].width = 26
    r = 4
    for name, total in sorted(by_cat.items(), key=lambda kv: -kv[1]):
        ws.cell(row=r, column=11, value=safe_excel_text(name))
        ws.cell(row=r, column=12, value=round(total, 2)).number_format = EUR_FMT
        r += 1
    cat_last = r - 1

    if last < 4:
        return  # no data, no charts

    # ---- charts ----
    bar = BarChart()
    bar.title = "Einnahmen vs. Ausgaben je Monat"
    bar.type = "col"
    bar.grouping = "clustered"
    data = Reference(ws, min_col=2, max_col=3, min_row=3, max_row=last)
    cats = Reference(ws, min_col=1, min_row=4, max_row=last)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.height, bar.width = 8, 16
    ws.add_chart(bar, "A" + str(last + 3))

    line = LineChart()
    line.title = "Kumuliertes Netto (cumulative net)"
    data = Reference(ws, min_col=6, min_row=3, max_row=last)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    line.height, line.width = 8, 16
    ws.add_chart(line, "A" + str(last + 20))

    if treat_last >= 4:
        pie = PieChart()
        pie.title = "Einnahmen nach Behandlung"
        data = Reference(ws, min_col=9, min_row=3, max_row=treat_last)
        labels = Reference(ws, min_col=8, min_row=4, max_row=treat_last)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height, pie.width = 8, 12
        ws.add_chart(pie, "J" + str(last + 3))

    if cat_last >= 4:
        pie2 = PieChart()
        pie2.title = "Ausgaben nach Kategorie"
        data = Reference(ws, min_col=12, min_row=3, max_row=cat_last)
        labels = Reference(ws, min_col=11, min_row=4, max_row=cat_last)
        pie2.add_data(data, titles_from_data=True)
        pie2.set_categories(labels)
        pie2.height, pie2.width = 8, 12
        ws.add_chart(pie2, "J" + str(last + 20))
