"""Migration tests (audit A-01 / P0-1): dry-run purity, apply + reconcile,
idempotency, conflict refusal, validation refusal, failure rollback.

A synthetic v1 source (workbook + photos + json files) is built in a tmp dir
with the EXACT column layouts of the original create-workbook.py. The real v1
folder is never touched — conftest.py aborts the run if isolation is broken.
"""
from __future__ import annotations

import hashlib
import io
import json
import sqlite3
import zipfile
from pathlib import Path

import pytest


@pytest.fixture(scope="session")
def migrate(data_dir):
    from app import migrate as m
    return m


def _sha(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _tiny_jpeg() -> bytes:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", (40, 40), "#c9a24b").save(buf, "JPEG")
    return buf.getvalue()


OUT_ROWS = [
    ["001", "2026-06-01", "dm Drogerie", "Salon Supplies", 12.5, "EUR", 12.5,
     "RC-1", "48.780000, 9.180000", "note 1", "Business", "OK",
     "01.06.2026 10:00", "01.06.2026 10:01"],
    ["002", "2026-06-02", "Migros", "General Business Expenses", 20.0, "CHF", None,
     "", "", "", "Business", "CHECK FX", "", ""],
]
IN_ROWS = [
    ["001", "2026-06-03", "Anna", "Microblading", 350.0, "EUR", 350.0,
     "", "", "", "OK", "", ""],
]
TREATMENTS = [{"name": "Microblading", "price": 250},
              {"name": "Lip Blush", "price": 300},
              {"name": "Brow Touch-up", "price": 90}]
PLACES = [{"name": "Zuhause", "address": "Hauptstr. 1"},
          {"name": "Studio", "address": "Freiburg"}]
OPEN_TRIP = {"purpose": "Client appointment", "start": "2026-07-01T09:00",
             "legs": [{"date": "2026-07-01", "from": "Zuhause", "to": "Basel",
                       "km": 75.0, "overnight": False}],
             "createdAt": "2026-07-01T09:00"}


def make_v1_source(root: Path, migrate) -> Path:
    """Build a faithful miniature v1 app folder."""
    from openpyxl import Workbook

    root.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws_out = wb.active
    ws_out.title = "Out"
    ws_out.append(migrate.OUT_HEADERS)
    for r in OUT_ROWS:
        ws_out.append(r)
    ws_in = wb.create_sheet("In")
    ws_in.append(migrate.IN_HEADERS)
    for r in IN_ROWS:
        ws_in.append(r)
    wb.save(root / "her-expenses.xlsx")

    photos = root / "her-photos"
    photos.mkdir()
    (photos / "O001.jpg").write_bytes(_tiny_jpeg())
    (photos / "I001.jpg").write_bytes(_tiny_jpeg())

    (root / "treatments.json").write_text(json.dumps(TREATMENTS), encoding="utf-8")
    (root / "places.json").write_text(json.dumps(PLACES), encoding="utf-8")
    (root / "open-trip.json").write_text(json.dumps(OPEN_TRIP), encoding="utf-8")
    return root


def _db(target: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(target / "sheila.db")
    conn.row_factory = sqlite3.Row
    return conn


# ---------------------------------------------------------------------------

def test_dry_run_reports_and_writes_nothing(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"
    before = {p.name: _sha(p) for p in src.rglob("*") if p.is_file()}

    report = migrate.run(src, data_dir=target, apply_changes=False)

    assert report["dry_run"] and not report["applied"]
    assert report["counts"] == {"out": 2, "in": 1, "photos": 2, "treatments": 3,
                                "places": 2, "open_trip": 1,
                                "eur_out": 12.5, "eur_in": 350.0}
    assert any("O002.jpg" in w for w in report["warnings"])  # row 2 has no photo
    assert not target.exists()  # dry run created nothing at the target
    after = {p.name: _sha(p) for p in src.rglob("*") if p.is_file()}
    assert before == after  # source untouched


def test_apply_migrates_and_reconciles(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"

    report = migrate.run(src, data_dir=target, apply_changes=True)
    assert report["applied"] and not report["already_migrated"]

    conn = _db(target)
    try:
        rows = conn.execute(
            "SELECT * FROM receipts ORDER BY kind DESC, serial").fetchall()
        assert [(r["kind"], r["serial"]) for r in rows] == [
            ("out", 1), ("out", 2), ("in", 1)]
        o1 = rows[0]
        assert (o1["date"], o1["merchant"], o1["article"]) == \
            ("2026-06-01", "dm Drogerie", "Salon Supplies")
        assert o1["eur"] == 12.5 and o1["status"] == "OK"
        assert o1["photo"] == "O001.jpg" and o1["generated_card"] == 0
        assert o1["location"] == "48.780000, 9.180000"
        assert o1["map_url"].startswith("https://www.google.com/maps?q=48.780000,9.180000")
        assert o1["saved_at"] == "01.06.2026 10:01"
        o2 = rows[1]
        assert o2["eur"] is None and o2["status"] == "CHECK FX" and o2["photo"] == ""
        i1 = rows[2]
        assert (i1["client_name"], i1["treatment"], i1["eur"]) == ("Anna", "Microblading", 350.0)
        assert i1["photo"] == "I001.jpg"

        treats = conn.execute("SELECT name, price FROM treatments ORDER BY pos").fetchall()
        assert [(t["name"], t["price"]) for t in treats] == \
            [("Microblading", 250.0), ("Lip Blush", 300.0), ("Brow Touch-up", 90.0)]

        places = conn.execute("SELECT name FROM places ORDER BY pos").fetchall()
        assert [p["name"] for p in places] == ["Zuhause", "Studio"]

        payload = conn.execute("SELECT payload FROM open_trip WHERE id=1").fetchone()
        assert json.loads(payload["payload"]) == OPEN_TRIP

        fp = conn.execute(
            "SELECT value FROM meta WHERE key='migration.source_fingerprint'").fetchone()
        assert fp["value"] == report["fingerprint"]
    finally:
        conn.close()

    # photos copied and byte-identical to the source
    assert sorted(p.name for p in (target / "photos").iterdir()) == ["I001.jpg", "O001.jpg"]
    assert _sha(target / "photos" / "O001.jpg") == _sha(src / "her-photos" / "O001.jpg")
    assert not list(target.glob(".migration-staging*"))  # no leftover staging dirs

    # verified backup exists and contains the workbook + photos
    backups = list((target / "backups").glob("pre-migration-*.zip"))
    assert len(backups) == 1
    with zipfile.ZipFile(backups[0]) as z:
        names = set(z.namelist())
    assert "her-expenses.xlsx" in names and "her-photos/O001.jpg" in names


def test_apply_is_idempotent(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"
    migrate.run(src, data_dir=target, apply_changes=True)

    report2 = migrate.run(src, data_dir=target, apply_changes=True)
    assert report2["already_migrated"] and not report2["applied"]

    conn = _db(target)
    try:
        assert conn.execute("SELECT COUNT(*) c FROM receipts").fetchone()["c"] == 3
        assert conn.execute("SELECT COUNT(*) c FROM treatments").fetchone()["c"] == 3
    finally:
        conn.close()
    # still exactly one backup — the no-op re-run wrote nothing
    assert len(list((target / "backups").glob("*.zip"))) == 1


def test_apply_refuses_changed_source(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"
    migrate.run(src, data_dir=target, apply_changes=True)

    (src / "treatments.json").write_text(
        json.dumps(TREATMENTS + [{"name": "New", "price": 1}]), encoding="utf-8")
    with pytest.raises(migrate.MigrationError) as e:
        migrate.run(src, data_dir=target, apply_changes=True)
    assert e.value.code == 3


def test_apply_refuses_nonempty_target(tmp_path, migrate):
    from app import db as appdb

    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"
    target.mkdir()
    conn = sqlite3.connect(target / "sheila.db")
    conn.row_factory = sqlite3.Row
    try:
        appdb.init_schema(conn)
        conn.execute("INSERT INTO receipts (kind, serial) VALUES ('out', 1)")
        conn.commit()
    finally:
        conn.close()

    with pytest.raises(migrate.MigrationError) as e:
        migrate.run(src, data_dir=target, apply_changes=True)
    assert e.value.code == 3


def test_apply_refuses_overlapping_dirs(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    with pytest.raises(migrate.MigrationError) as e:
        migrate.run(src, data_dir=src, apply_changes=True)
    assert e.value.code == 3


def test_header_mismatch_refused(tmp_path, migrate):
    from openpyxl import load_workbook

    src = make_v1_source(tmp_path / "v1", migrate)
    wb = load_workbook(src / "her-expenses.xlsx")
    wb["Out"]["C1"] = "Vendor"  # renamed header must be rejected
    wb.save(src / "her-expenses.xlsx")

    with pytest.raises(migrate.MigrationError) as e:
        migrate.run(src, data_dir=tmp_path / "target", apply_changes=False)
    assert e.value.code == 2 and "header mismatch" in str(e.value)


@pytest.mark.parametrize("mutate", ["treatments.json", "workbook", "photo"])
def test_source_mutated_between_read_and_backup_aborts(tmp_path, migrate, mutate):
    """TOCTOU guard: the backup is verified against READ-time hashes, so a
    source file changing after read_source() must abort the whole migration —
    no rows, no fingerprint, no photos, and no mismatched backup left behind."""
    src_dir = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"

    srcdata = migrate.read_source(src_dir)  # state A is read and hashed here

    # ... then the source changes to state B before/while the backup is taken
    if mutate == "treatments.json":
        (src_dir / "treatments.json").write_text(
            json.dumps(TREATMENTS + [{"name": "Sneaky", "price": 1}]), encoding="utf-8")
    elif mutate == "workbook":
        from openpyxl import load_workbook

        wb = load_workbook(src_dir / "her-expenses.xlsx")
        wb["Out"].append(["003", "2026-06-09", "Late Shop", "Fuel", 5.0, "EUR", 5.0,
                          "", "", "", "Business", "OK", "", ""])
        wb.save(src_dir / "her-expenses.xlsx")
    else:
        (src_dir / "her-photos" / "O001.jpg").write_bytes(b"not the same jpeg anymore")

    with pytest.raises(migrate.MigrationError, match="changed on disk"):
        migrate.apply_migration(srcdata, target)

    conn = _db(target)
    try:
        assert conn.execute("SELECT COUNT(*) c FROM receipts").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) c FROM treatments").fetchone()["c"] == 0
        assert conn.execute(
            "SELECT COUNT(*) c FROM meta WHERE key LIKE 'migration.%'").fetchone()["c"] == 0
    finally:
        conn.close()
    photos = target / "photos"
    assert not photos.exists() or not any(photos.iterdir())
    assert not list((target / "backups").glob("*.zip"))  # bad backup was removed
    assert not list(target.glob(".migration-staging*"))


def test_staging_never_deletes_foreign_directories(tmp_path, migrate):
    """A pre-existing '.migration-staging' path (e.g. user debris) must survive:
    each run stages into its own unique mkdtemp directory."""
    src_dir = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"
    target.mkdir()
    foreign = target / ".migration-staging"
    foreign.mkdir()
    sentinel = foreign / "precious.txt"
    sentinel.write_text("do not delete me", encoding="utf-8")

    report = migrate.run(src_dir, data_dir=target, apply_changes=True)
    assert report["applied"]
    assert sentinel.read_text(encoding="utf-8") == "do not delete me"
    # and the run's own unique staging dir is gone
    assert [p.name for p in target.glob(".migration-staging*")] == [".migration-staging"]


def test_failure_during_photo_promotion_rolls_back(tmp_path, migrate, monkeypatch):
    src = make_v1_source(tmp_path / "v1", migrate)
    target = tmp_path / "target"

    real = migrate._sha256_file

    def corrupt_final_hashes(path):
        # Simulate corruption detected in the final photo location only.
        if Path(path).parent.name == "photos":
            return "0" * 64
        return real(path)

    monkeypatch.setattr(migrate, "_sha256_file", corrupt_final_hashes)
    with pytest.raises(migrate.MigrationError, match="hash verification"):
        migrate.run(src, data_dir=target, apply_changes=True)

    conn = _db(target)
    try:
        assert conn.execute("SELECT COUNT(*) c FROM receipts").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) c FROM treatments").fetchone()["c"] == 0
        assert conn.execute("SELECT COUNT(*) c FROM open_trip").fetchone()["c"] == 0
        assert conn.execute(
            "SELECT COUNT(*) c FROM meta WHERE key LIKE 'migration.%'").fetchone()["c"] == 0
    finally:
        conn.close()
    assert list((target / "photos").iterdir()) == []  # promoted file removed
    assert not list(target.glob(".migration-staging*"))  # no leftover staging dirs

    # after the fault is gone, the same apply succeeds cleanly
    monkeypatch.setattr(migrate, "_sha256_file", real)
    report = migrate.run(src, data_dir=target, apply_changes=True)
    assert report["applied"]


def test_cli_exit_codes(tmp_path, migrate):
    src = make_v1_source(tmp_path / "v1", migrate)
    assert migrate.main(["--source", str(src),
                         "--data-dir", str(tmp_path / "t1")]) == 0  # dry run
    assert migrate.main(["--source", str(src),
                         "--data-dir", str(tmp_path / "t1"), "--apply"]) == 0
    assert migrate.main(["--source", str(tmp_path / "nowhere"),
                         "--data-dir", str(tmp_path / "t2")]) == 2  # bad source
