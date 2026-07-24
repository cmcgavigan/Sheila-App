"""Isolated port of the old tests/e2e.py flow (audit A-02 / P0-2).

Same API drive: receipts (manual + photo), FX flagging, treatments PIN, a full
multi-leg trip, places, export — then the workbook is opened with openpyxl and
its rows/formulas/classification/charts are checked.

The tests in this module form ONE sequential flow against a shared isolated
database (serial numbers, trip state and export contents build on each other),
exactly like the script they replace. pytest runs them in definition order —
keep it that way.

Honesty note (audit A-13): openpyxl checks formula STRINGS and stored values.
Nothing here recalculates formulas in Excel, drives a browser, or tests the
offline queue. Those remain open items for P4-1.
"""
from __future__ import annotations

import base64
import io
import os

WRONG_PIN = "0000"


def _pin() -> str:
    # conftest sets TREATMENTS_PIN before app.config is imported
    return os.environ["TREATMENTS_PIN"]


def _jpeg_b64(size=(800, 1000), color="#888") -> str:
    from PIL import Image

    buf = io.BytesIO()
    Image.new("RGB", size, color).save(buf, "JPEG")
    return base64.b64encode(buf.getvalue()).decode()


# --- health / shell ---------------------------------------------------------

def test_health(client):
    r = client.get("/api/health")
    assert r.status_code == 200
    body = r.json()
    assert body["ok"] is True
    assert body["receipts"] == {"out": 0, "in": 0}


def test_pwa_shell_served(client):
    r = client.get("/")
    assert r.status_code == 200 and "sheila" in r.text.lower()
    r = client.get("/sw.js")
    assert r.status_code == 200 and "sheila-shell-v9" in r.text
    r = client.get("/setup")
    assert r.status_code == 200 and "data:image/png;base64" in r.text
    r = client.get("/export")
    assert r.status_code == 200


# --- income -----------------------------------------------------------------

def test_save_in_manual_generates_card(client):
    r = client.post("/api/save-in", json={
        "clientName": "Test Client", "treatment": "Microblading",
        "date": "2026-07-20", "totalCost": 350, "currency": "EUR",
        "receiptCode": "INV-1", "note": "e2e",
    })
    assert r.status_code == 200
    j = r.json()
    assert j["ok"] and j["serial"] == 1
    assert j["generated"] is True and j["eur"] == 350


def test_save_in_chf_flagged_without_fx(client, monkeypatch):
    # Isolated tests must not hit frankfurter.app; simulate FX being
    # unavailable — the row must be flagged CHECK FX with no EUR value.
    from app import services

    async def _no_fx(total, currency, receipt_date):
        return {"eur": None, "rate": None, "rate_date": None, "flagged": True}

    monkeypatch.setattr(services, "to_eur", _no_fx)
    r = client.post("/api/save-in", json={
        "clientName": "Swiss Client", "treatment": "Lip Blush",
        "date": "2026-07-10", "totalCost": 400, "currency": "CHF",
    })
    assert r.status_code == 200
    j = r.json()
    assert j["ok"] and j["serial"] == 2
    assert j["eur"] is None and j["fxFlagged"] is True


# --- expenses ---------------------------------------------------------------

def test_save_out_with_photo(client, data_dir):
    r = client.post("/api/save-out", json={
        "merchant": "dm Drogerie", "article": "Salon Supplies",
        "date": "2026-07-21", "totalCost": 23.45, "currency": "EUR",
        "image": "data:image/jpeg;base64," + _jpeg_b64(),
        "businessPersonal": "Business", "lat": 48.78, "lng": 9.18,
    })
    assert r.status_code == 200
    j = r.json()
    assert j["ok"] and j["serial"] == 1 and j["generated"] is False
    assert (data_dir / "photos" / j["photo"]).is_file()


def test_save_out_bewirtung(client):
    r = client.post("/api/save-out", json={
        "merchant": "Restaurant Adler", "article": "Client Entertainment",
        "date": "2026-07-22", "totalCost": 80, "currency": "EUR",
        "businessPersonal": "Business",
    })
    assert r.status_code == 200 and r.json()["serial"] == 2


def test_save_out_personal(client):
    r = client.post("/api/save-out", json={
        "merchant": "Zara", "article": "General Business Expenses",
        "date": "2026-07-22", "totalCost": 50, "currency": "EUR",
        "businessPersonal": "Personal",
    })
    assert r.status_code == 200 and r.json()["serial"] == 3


def test_save_out_fuel(client):
    r = client.post("/api/save-out", json={
        "merchant": "Shell", "article": "Fuel", "date": "2026-07-23",
        "totalCost": 65, "currency": "EUR", "businessPersonal": "Business",
    })
    assert r.status_code == 200 and r.json()["serial"] == 4


# --- treatments PIN gate ----------------------------------------------------

def test_treatments_wrong_pin_rejected(client):
    r = client.post("/api/treatments", json={"pin": WRONG_PIN, "treatments": []})
    assert r.status_code == 403


def test_treatments_saved_and_read_back(client):
    r = client.post("/api/treatments", json={
        "pin": _pin(),
        "treatments": [{"name": "Microblading", "price": 350},
                       {"name": "Lip Blush", "price": 400}],
    })
    assert r.status_code == 200 and len(r.json()["treatments"]) == 2
    r = client.get("/api/treatments")
    assert [t["name"] for t in r.json()["treatments"]] == ["Microblading", "Lip Blush"]


# --- trip lifecycle: 3 legs, 2 overnights -----------------------------------

def test_trip_lifecycle_multi_leg(client):
    r = client.post("/api/trip/end", json={})
    assert r.status_code == 409  # no open trip yet

    r = client.post("/api/trip/start", json={"purpose": "Training / course"})
    assert r.status_code == 200 and r.json()["ok"]
    r = client.post("/api/trip/start", json={"purpose": "Client appointment"})
    assert r.status_code == 409  # second open trip blocked

    for a, b, km, overnight in [("Zuhause", "Stuttgart", 220.0, True),
                                ("Stuttgart", "Frankfurt", 210.0, True),
                                ("Frankfurt", "Zuhause", 220.0, False)]:
        r = client.post("/api/trip/leg",
                        json={"from": a, "to": b, "km": km, "overnight": overnight})
        assert r.status_code == 200 and r.json()["ok"]

    r = client.get("/api/trip/active")
    assert len(r.json()["trip"]["legs"]) == 3

    r = client.post("/api/trip/end", json={
        "mealReduction": 11.20, "overnightCost": 180, "incidentals": 15,
        "note": "e2e trip",
    })
    assert r.status_code == 200
    j = r.json()
    assert j["ok"] and j["km"] == 650.0 and j["nights"] == 2 and j["legs"] == 3

    r = client.get("/api/trip/active")
    assert r.json()["trip"] is None


def test_trip_single_day(client):
    client.post("/api/trip/start", json={"purpose": "Client appointment"})
    client.post("/api/trip/leg",
                json={"from": "Zuhause", "to": "Karlsruhe", "km": 80, "overnight": False})
    client.post("/api/trip/leg",
                json={"from": "Karlsruhe", "to": "Zuhause", "km": 80, "overnight": False})
    r = client.post("/api/trip/end", json={})
    assert r.status_code == 200 and r.json()["km"] == 160.0


# --- places -----------------------------------------------------------------

def test_places_saved(client):
    r = client.post("/api/places", json={"places": [
        {"name": "Zuhause", "address": "Hauptstr. 1, 79098"},
        {"name": "Studio", "address": "Freiburg"},
    ]})
    assert r.status_code == 200 and len(r.json()["places"]) == 2


# --- export + workbook verification -----------------------------------------

def test_export_workbook(client):
    from openpyxl import load_workbook

    r = client.post("/api/export")
    assert r.status_code == 200
    j = r.json()
    assert j["ok"]
    dl = client.get(j["download"])
    assert dl.status_code == 200 and len(dl.content) > 5000

    wb = load_workbook(io.BytesIO(dl.content))
    assert wb.sheetnames == ["Info", "Reisen", "Fahrten", "Ausgaben", "Einnahmen",
                             "Dashboard"]

    info = wb["Info"]
    assert info["B6"].value == 0.30  # km rate

    aus = wb["Ausgaben"]
    rows = {aus.cell(row=r_, column=3).value: r_ for r_ in range(2, 8)}
    r_bew = rows.get("Restaurant Adler")
    assert r_bew is not None
    assert "VLOOKUP" in str(aus.cell(row=r_bew, column=8).value)
    r_pers = rows.get("Zara")
    assert r_pers is not None and aus.cell(row=r_pers, column=8).value == 0
    assert aus.cell(row=r_pers, column=10).value == "Privat"
    r_fuel = rows.get("Shell")
    assert r_fuel is not None and aus.cell(row=r_fuel, column=1).hyperlink is not None

    reisen = wb["Reisen"]
    assert reisen["H2"].value == 650.0
    assert str(reisen["I2"].value) == "=H2*Info!$B$6"
    assert "IF(F2>0" in str(reisen["J2"].value)  # per-diem references nights
    assert str(reisen["N2"].value) == "=I2+J2-K2+L2+M2"
    # expected trip-1 maths (formula strings above, not recalculated here):
    #   mileage 650 * 0.30 = 195.00 ; per-diem 2 nights * 28 = 56.00
    #   total = 195 + 56 - 11.20 + 180 + 15 = 434.80

    fahrten = wb["Fahrten"]
    legs_rows = [fahrten.cell(row=r_, column=1).value for r_ in range(2, 8)]
    assert legs_rows[:5] == [1, 1, 1, 2, 2]

    ein = wb["Einnahmen"]
    clients = {ein.cell(row=r_, column=3).value for r_ in (2, 3)}
    assert clients == {"Test Client", "Swiss Client"}
    # The CHF row was saved with FX unavailable: EUR empty + status CHECK FX.
    r_chf = 2 if ein["C2"].value == "Swiss Client" else 3
    assert ein.cell(row=r_chf, column=7).value is None
    assert ein.cell(row=r_chf, column=9).value == "CHECK FX"

    dash = wb["Dashboard"]
    assert len(dash._charts) >= 3
