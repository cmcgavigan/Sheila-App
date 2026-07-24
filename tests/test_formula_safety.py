"""Excel formula-injection regression coverage."""
from __future__ import annotations

from openpyxl import load_workbook


def test_safe_excel_text_neutralizes_formula_prefixes():
    from app.export.steuer import safe_excel_text

    for prefix in "=+-@\t\r\n":
        value = prefix + "payload"
        assert safe_excel_text(value) == "'" + value
    assert safe_excel_text("ordinary") == "ordinary"
    assert safe_excel_text(12.5) == 12.5


def test_export_writes_user_text_as_literal_and_keeps_own_formulas(tmp_path, data_dir):
    from app import db
    from app.export.steuer import generate_workbook

    db.init_db()
    operation_ids = ("formula-out", "formula-in")
    with db.get_db() as conn:
        db.insert_receipt(conn, {
            "kind": "out", "serial": db.next_serial(conn, "out"),
            "merchant": "=HYPERLINK(\"https://evil\")",
            "article": "+CMD", "total": 1, "currency": "EUR", "eur": 1,
            "business_personal": "Business", "receipt_code": "@ref",
            "status": "OK", "note": "-malicious", "operation_id": operation_ids[0],
            "operation_hash": "formula-test-out",
        })
        db.insert_receipt(conn, {
            "kind": "in", "serial": db.next_serial(conn, "in"),
            "client_name": "=SUM(1,1)", "treatment": "@service", "total": 2,
            "currency": "EUR", "eur": 2, "receipt_code": "\tref",
            "status": "OK", "note": "ordinary", "operation_id": operation_ids[1],
            "operation_hash": "formula-test-in",
        })
    try:
        path = generate_workbook(tmp_path / "safe.xlsx")
        wb = load_workbook(path, data_only=False)
        out = wb["Ausgaben"]
        assert any(out.cell(r, 3).value == "'=HYPERLINK(\"https://evil\")"
                   for r in range(2, out.max_row + 1))
        assert any(out.cell(r, 4).value == "'+CMD"
                   for r in range(2, out.max_row + 1))
        assert any(out.cell(r, 9).data_type == "f" and
                   str(out.cell(r, 9).value).startswith("=IF(")
                   for r in range(2, out.max_row + 1))
        inc = wb["Einnahmen"]
        assert any(inc.cell(r, 3).value == "'=SUM(1,1)"
                   for r in range(2, inc.max_row + 1))
    finally:
        with db.get_db() as conn:
            conn.execute(
                "DELETE FROM receipts WHERE operation_id IN (?, ?)", operation_ids
            )


def test_default_exports_are_unique_and_leave_no_temp_files(data_dir):
    from app import db
    from app.export.steuer import generate_workbook

    db.init_db()
    first = generate_workbook()
    second = generate_workbook()
    try:
        assert first != second
        assert first.is_file() and second.is_file()
        assert not list(data_dir.joinpath("exports").glob(".*.tmp"))
    finally:
        first.unlink(missing_ok=True)
        second.unlink(missing_ok=True)
