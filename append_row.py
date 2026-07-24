#!/usr/bin/env python3
"""
append_row.py — append ONE row to the Out or In sheet of her-expenses.xlsx,
preserving the Dashboard charts and all live formulas.

Why Python/openpyxl and not the Node SheetJS lib:
  SheetJS rebuilds the whole workbook on write and DESTROYS charts + formulas.
  openpyxl opens the existing file and appends in place, leaving charts/formulas
  untouched. This script is spawned by server.js on every save.

Usage:
  python append_row.py --sheet out --excel /path/to/her-expenses.xlsx
  (row payload is read as JSON from stdin)

Prints a JSON result to stdout: {"ok": true, "serial": "001", "rowCount": N}

Column layouts (must match create-workbook.py):

  Out:  # | Date | Merchant | Article | Total | Currency | EUR Value | Receipt Code |
        Location | Note | Business/Personal | Status | Captured At | Saved At
  In:   # | Date | Client Name | Treatment | Total | Currency | EUR Value | Receipt Code |
        Location | Note | Status | Captured At | Saved At

Data starts at row 2 (row 1 = header). Serials are 3-digit, sequential, per sheet.
"""
import sys, json, argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def col_count_out():
    return 14


def build_out_values(serial, r):
    # Order matches the Out header exactly.
    eur = r.get("eur", None)
    return [
        serial,                                   # #
        r.get("date", ""),                        # Date
        r.get("merchant", ""),                    # Merchant
        r.get("article", ""),                     # Article
        float(r.get("total") or 0),               # Total
        r.get("currency", ""),                    # Currency
        (None if eur is None else float(eur)),    # EUR Value (blank if FX failed)
        r.get("receiptCode", ""),                 # Receipt Code
        r.get("location", ""),                    # Location
        r.get("note", ""),                        # Note
        r.get("businessPersonal", "Business"),    # Business/Personal
        r.get("status", "OK"),                    # Status
        r.get("capturedAt", ""),                  # Captured At
        r.get("savedAt", ""),                     # Saved At
    ]


def build_in_values(serial, r):
    eur = r.get("eur", None)
    return [
        serial,                                   # #
        r.get("date", ""),                        # Date
        r.get("clientName", ""),                  # Client Name
        r.get("treatment", ""),                   # Treatment
        float(r.get("total") or 0),               # Total
        r.get("currency", ""),                    # Currency
        (None if eur is None else float(eur)),    # EUR Value
        r.get("receiptCode", ""),                 # Receipt Code
        r.get("location", ""),                    # Location
        r.get("note", ""),                        # Note
        r.get("status", "OK"),                    # Status
        r.get("capturedAt", ""),                  # Captured At
        r.get("savedAt", ""),                     # Saved At
    ]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheet", required=True, choices=["out", "in"])
    ap.add_argument("--excel", required=True)
    args = ap.parse_args()

    row = json.loads(sys.stdin.read() or "{}")
    sheet_name = "Out" if args.sheet == "out" else "In"

    wb = load_workbook(args.excel)  # keeps charts + formulas
    if sheet_name not in wb.sheetnames:
        print(json.dumps({"ok": False, "error": f"Sheet {sheet_name} missing"}))
        sys.exit(1)
    ws = wb[sheet_name]

    # Next data row = first empty row at/after row 2. ws.max_row counts the header
    # even on an empty sheet, so compute the real last populated data row.
    last = 1
    for rr in range(ws.max_row, 1, -1):
        if ws.cell(row=rr, column=1).value not in (None, ""):
            last = rr
            break
    new_row = last + 1
    serial_int = new_row - 1  # row 2 -> serial 1
    serial = f"{serial_int:03d}"

    values = build_out_values(serial, row) if args.sheet == "out" else build_in_values(serial, row)
    for ci, val in enumerate(values, start=1):
        ws.cell(row=new_row, column=ci, value=val)

    # Hyperlink the # cell to the receipt photo (only if there IS a photo —
    # manual entries have none), and the Location cell to Google Maps.
    prefix = row.get("photoPrefix", "O" if args.sheet == "out" else "I")
    photos_dir = "her-photos"
    if row.get("hasPhoto", True):
        a_cell = ws.cell(row=new_row, column=1)
        a_cell.hyperlink = f"{photos_dir}/{prefix}{serial}.jpg"
        a_cell.style = "Hyperlink"

    # Location column index differs per sheet (Out col 9, In col 9 — both 9).
    loc_col = 9
    map_url = row.get("mapUrl", "")
    if map_url:
        loc_cell = ws.cell(row=new_row, column=loc_col)
        loc_cell.hyperlink = map_url
        loc_cell.style = "Hyperlink"

    wb.save(args.excel)

    data_rows = new_row - 1  # excluding header
    print(json.dumps({"ok": True, "serial": serial, "rowCount": data_rows}))


if __name__ == "__main__":
    main()
