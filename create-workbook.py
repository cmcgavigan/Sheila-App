#!/usr/bin/env python3
"""
create-workbook.py — ONE-TIME builder for her-expenses.xlsx.

Builds the full template: Out, In, Combined and Dashboard sheets with headers,
live formulas (SUMIF/SUMIFS), frozen totals, and native charts. The server then
APPENDS rows via append_row.py without ever recreating this structure.

Run once:   python create-workbook.py
It refuses to overwrite an existing workbook (rename/move the old one first) so a
re-run can never wipe real data.

=========================  v1 / REVISIT PLACEHOLDER  =========================
The Excel + charts layer is the agreed "build now, polish later" piece. The
Dashboard formulas reference a generous fixed range (rows 2..MAXROWS) on Out/In
so newly appended rows are picked up automatically without touching this file.
Charts bind to helper tables on the Dashboard that are themselves formula-driven.
When revisiting: tune chart styling, add any missing KPI, or swap the FX engine.
=============================================================================
"""
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, DoughnutChart, Reference
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.environ.get("EXCEL_PATH", "her-expenses.xlsx")
if not os.path.isabs(EXCEL_PATH):
    EXCEL_PATH = os.path.join(HERE, EXCEL_PATH)

# Generous data window the dashboard formulas scan. Appends land inside this.
MAXROWS = 5000

# Number of treatment rows in the Dashboard's Income-by-Treatment table.
TREATMENT_SLOTS = 18


def load_treatment_names():
    """Read the treatment names from treatments.json so the Dashboard's
    Income-by-Treatment table totals against the real treatments she's set up.
    Falls back to blank slots if the file is missing/unreadable."""
    import json
    path = os.environ.get("TREATMENTS_PATH", os.path.join(HERE, "treatments.json"))
    if not os.path.isabs(path):
        path = os.path.join(HERE, path)
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        names = [str(t.get("name", "")).strip() for t in data if str(t.get("name", "")).strip()]
        # de-dupe, preserve order
        seen, out = set(), []
        for n in names:
            if n not in seen:
                seen.add(n); out.append(n)
        return out[:TREATMENT_SLOTS]
    except Exception as e:
        print(f"  (treatments.json not loaded for dashboard: {e})", file=sys.stderr)
        return []

ARTICLES = [
    "Fuel", "Meals (Travel)", "Client Entertainment", "Accommodation", "Parking",
    "Vehicle expenses", "Phone & Internet", "Insurance", "Training & Education",
    "Marketing & Advertising", "Equipment & Tools", "General Business Expenses",
    "PMU Pigments & Inks", "PMU Needles & Cartridges", "PMU Equipment & Machines",
    "Skincare & Treatment Products", "Disposables & PPE", "Salon Supplies",
]

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ---- styling helpers ----
PINK = "B6356B"
PINK_LT = "F8D7E6"
INK = "1F2430"
GREY_LT = "F2F2F4"
THIN = Side(style="thin", color="D9D9DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_row(ws, headers, row=1):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=PINK)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 28


def title_cell(ws, addr, text, size=14):
    c = ws[addr]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=size, color=PINK)


def label(ws, addr, text, bold=True):
    c = ws[addr]
    c.value = text
    c.font = Font(name="Arial", bold=bold, size=11, color=INK)


def kpi(ws, addr, formula, numfmt="#,##0.00"):
    c = ws[addr]
    c.value = formula
    c.font = Font(name="Arial", bold=True, size=12, color=INK)
    c.number_format = numfmt


# =====================================================================
# Build
# =====================================================================
def build():
    if os.path.exists(EXCEL_PATH):
        print(f"[!] {EXCEL_PATH} already exists — refusing to overwrite. "
              f"Move/rename it first if you really want a fresh template.", file=sys.stderr)
        sys.exit(2)

    wb = Workbook()

    # ---------------- Sheet 1: Out (expenses) ----------------
    ws_out = wb.active
    ws_out.title = "Out"
    out_headers = ["#", "Date", "Merchant", "Article", "Total", "Currency", "EUR Value",
                   "Receipt Code", "Location", "Note", "Business/Personal", "Status",
                   "Captured At", "Saved At"]
    header_row(ws_out, out_headers)
    ws_out.freeze_panes = "A2"
    widths = [6, 12, 22, 24, 10, 9, 11, 18, 24, 22, 16, 10, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws_out.column_dimensions[get_column_letter(i)].width = w
    # EUR column number format
    for r in range(2, MAXROWS + 1):
        ws_out.cell(row=r, column=5).number_format = "#,##0.00"
        ws_out.cell(row=r, column=7).number_format = "#,##0.00"

    # ---------------- Sheet 2: In (income) ----------------
    ws_in = wb.create_sheet("In")
    in_headers = ["#", "Date", "Client Name", "Treatment", "Total", "Currency", "EUR Value",
                  "Receipt Code", "Location", "Note", "Status", "Captured At", "Saved At"]
    header_row(ws_in, in_headers)
    ws_in.freeze_panes = "A2"
    widths_in = [6, 12, 22, 24, 10, 9, 11, 18, 24, 22, 10, 18, 18]
    for i, w in enumerate(widths_in, start=1):
        ws_in.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, MAXROWS + 1):
        ws_in.cell(row=r, column=5).number_format = "#,##0.00"
        ws_in.cell(row=r, column=7).number_format = "#,##0.00"

    # Out EUR col = G, In EUR col = G. Out Date = B, In Date = B.
    OUT_EUR = f"Out!$G$2:$G${MAXROWS}"
    OUT_DATE = f"Out!$B$2:$B${MAXROWS}"
    OUT_ART = f"Out!$D$2:$D${MAXROWS}"
    OUT_BP = f"Out!$K$2:$K${MAXROWS}"
    OUT_CUR = f"Out!$F$2:$F${MAXROWS}"
    OUT_TOTAL = f"Out!$E$2:$E${MAXROWS}"
    IN_EUR = f"In!$G$2:$G${MAXROWS}"
    IN_DATE = f"In!$B$2:$B${MAXROWS}"
    IN_TREAT = f"In!$D$2:$D${MAXROWS}"
    IN_CLIENT = f"In!$C$2:$C${MAXROWS}"

    # ---------------- Sheet 3: Combined ----------------
    # Row 1: frozen totals. Data mirrors Out then In via formulas.
    ws_c = wb.create_sheet("Combined")
    title_cell(ws_c, "A1", "Total In (EUR)")
    ws_c["B1"] = f"=SUM({IN_EUR})"; ws_c["B1"].number_format = "#,##0.00"; ws_c["B1"].font = Font(bold=True, color="0A7D34")
    title_cell(ws_c, "C1", "Total Out (EUR)")
    ws_c["D1"] = f"=SUM({OUT_EUR})"; ws_c["D1"].number_format = "#,##0.00"; ws_c["D1"].font = Font(bold=True, color="B00020")
    title_cell(ws_c, "E1", "Net (EUR)")
    ws_c["F1"] = "=B1-D1"; ws_c["F1"].number_format = "#,##0.00"; ws_c["F1"].font = Font(bold=True, color=INK)

    comb_headers = ["#", "Date", "Type", "Total", "Currency", "EUR Value"]
    header_row(ws_c, comb_headers, row=2)
    ws_c.freeze_panes = "A3"
    for i, w in enumerate([10, 12, 8, 10, 9, 11], start=1):
        ws_c.column_dimensions[get_column_letter(i)].width = w

    # Data starts row 3. First half mirrors Out rows, second half mirrors In rows.
    # Each Combined row links back to its source row. We lay Out block then In block.
    HALF = MAXROWS - 1  # rows of capacity per side
    # Out block: rows 3 .. 3+HALF-1
    for k in range(HALF):
        src = 2 + k                 # Out data row
        dst = 3 + k                 # Combined row
        # only show if the source has a serial
        ws_c.cell(row=dst, column=1, value=f'=IF(Out!A{src}="","",HYPERLINK("#Out!A{src}",Out!A{src}))')
        ws_c.cell(row=dst, column=2, value=f'=IF(Out!A{src}="","",Out!B{src})')
        ws_c.cell(row=dst, column=3, value=f'=IF(Out!A{src}="","","Out")')
        ws_c.cell(row=dst, column=4, value=f'=IF(Out!A{src}="","",Out!E{src})')
        ws_c.cell(row=dst, column=5, value=f'=IF(Out!A{src}="","",Out!F{src})')
        ws_c.cell(row=dst, column=6, value=f'=IF(Out!A{src}="","",Out!G{src})')
    in_start = 3 + HALF
    for k in range(HALF):
        src = 2 + k
        dst = in_start + k
        ws_c.cell(row=dst, column=1, value=f'=IF(In!A{src}="","",HYPERLINK("#In!A{src}",In!A{src}))')
        ws_c.cell(row=dst, column=2, value=f'=IF(In!A{src}="","",In!B{src})')
        ws_c.cell(row=dst, column=3, value=f'=IF(In!A{src}="","","In")')
        ws_c.cell(row=dst, column=4, value=f'=IF(In!A{src}="","",In!E{src})')
        ws_c.cell(row=dst, column=5, value=f'=IF(In!A{src}="","",In!F{src})')
        ws_c.cell(row=dst, column=6, value=f'=IF(In!A{src}="","",In!G{src})')

    # ---------------- Sheet 4: Dashboard ----------------
    ws_d = wb.create_sheet("Dashboard")
    ws_d.sheet_view.showGridLines = False
    ws_d.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws_d.column_dimensions[col].width = 15
    for col in "HIJKLMN":
        ws_d.column_dimensions[col].width = 13

    title_cell(ws_d, "A1", "Sheila McGavigan PMU — Dashboard", size=16)

    # Helper: month-of-year SUMPRODUCT over a date range + value range.
    # Dates are stored as text "YYYY-MM-DD", so we match on the "YYYY-MM" prefix.
    def month_sum(date_rng, val_rng, year_cell, month_num):
        mm = f"{month_num:02d}"
        # TEXT prefix match: LEFT(date,7) = year&"-"&mm
        return (f'=SUMPRODUCT(--(LEFT({date_rng},7)=($'
                f'{year_cell}&"-{mm}")),N(+{val_rng}))')

    # ---- Section 1: This Month ----
    title_cell(ws_d, "A3", "This Month", size=13)
    label(ws_d, "A4", "Month (YYYY-MM)")
    ws_d["B4"] = '=TEXT(TODAY(),"YYYY-MM")'
    ws_d["B4"].font = Font(bold=True, color="0000FF")  # blue input-ish (auto today)
    label(ws_d, "A5", "Total In (EUR)")
    ws_d["B5"] = f'=SUMPRODUCT(--(LEFT({IN_DATE},7)=$B$4),N(+{IN_EUR}))'; ws_d["B5"].number_format = "#,##0.00"
    label(ws_d, "A6", "Total Out (EUR)")
    ws_d["B6"] = f'=SUMPRODUCT(--(LEFT({OUT_DATE},7)=$B$4),N(+{OUT_EUR}))'; ws_d["B6"].number_format = "#,##0.00"
    label(ws_d, "A7", "Net (EUR)")
    ws_d["B7"] = "=B5-B6"; ws_d["B7"].number_format = "#,##0.00"
    label(ws_d, "A8", "Clients this month")
    ws_d["B8"] = f'=SUMPRODUCT(--(LEFT({IN_DATE},7)=$B$4),--({IN_CLIENT}<>""))'
    label(ws_d, "A9", "Avg transaction In (EUR)")
    ws_d["B9"] = "=IF(B8=0,0,B5/B8)"; ws_d["B9"].number_format = "#,##0.00"
    label(ws_d, "A10", "Biggest expense category")
    # INDEX/MATCH on the YTD article table built in Section 3 (H/I columns).
    ws_d["B10"] = '=IF(SUM($J$4:$J$21)=0,"—",INDEX($H$4:$H$21,MATCH(MAX($J$4:$J$21),$J$4:$J$21,0)))'

    # ---- Section 2: This Year, month by month ----
    title_cell(ws_d, "A13", "This Year — month by month", size=13)
    label(ws_d, "A14", "Year")
    ws_d["B14"] = '=TEXT(TODAY(),"YYYY")'
    ws_d["B14"].font = Font(bold=True, color="0000FF")
    yr_hdr = ["Month", "In (EUR)", "Out (EUR)", "Net (EUR)"]
    for ci, h in enumerate(yr_hdr, start=1):
        c = ws_d.cell(row=15, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=PINK)
        c.alignment = Alignment(horizontal="center")
    for mi, mname in enumerate(MONTHS):
        r = 16 + mi
        ws_d.cell(row=r, column=1, value=mname)
        mm = f"{mi+1:02d}"
        ws_d.cell(row=r, column=2,
                  value=f'=SUMPRODUCT(--(LEFT({IN_DATE},7)=($B$14&"-{mm}")),N(+{IN_EUR}))').number_format = "#,##0.00"
        ws_d.cell(row=r, column=3,
                  value=f'=SUMPRODUCT(--(LEFT({OUT_DATE},7)=($B$14&"-{mm}")),N(+{OUT_EUR}))').number_format = "#,##0.00"
        ws_d.cell(row=r, column=4, value=f"=B{r}-C{r}").number_format = "#,##0.00"
    # Annual totals row
    tr = 28
    ws_d.cell(row=tr, column=1, value="Annual total").font = Font(bold=True, color=PINK)
    ws_d.cell(row=tr, column=2, value="=SUM(B16:B27)").number_format = "#,##0.00"
    ws_d.cell(row=tr, column=3, value="=SUM(C16:C27)").number_format = "#,##0.00"
    ws_d.cell(row=tr, column=4, value="=SUM(D16:D27)").number_format = "#,##0.00"
    for ci in range(1, 5):
        ws_d.cell(row=tr, column=ci).font = Font(bold=True, color=PINK)

    # ---- Section 3: Expenses by Article (helper table H:J), this month + YTD ----
    title_cell(ws_d, "H2", "Expenses by Article", size=13)
    for ci, h in enumerate(["Article", "This Month", "YTD"], start=8):
        c = ws_d.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=PINK)
    for ai, art in enumerate(ARTICLES):
        r = 4 + ai
        ws_d.cell(row=r, column=8, value=art)
        # This month: article match AND month prefix match
        ws_d.cell(row=r, column=9,
                  value=f'=SUMPRODUCT(--({OUT_ART}=$H{r}),--(LEFT({OUT_DATE},7)=$B$4),N(+{OUT_EUR}))'
                  ).number_format = "#,##0.00"
        # YTD: article match AND year prefix match
        ws_d.cell(row=r, column=10,
                  value=f'=SUMPRODUCT(--({OUT_ART}=$H{r}),--(LEFT({OUT_DATE},4)=$B$14),N(+{OUT_EUR}))'
                  ).number_format = "#,##0.00"

    # ---- Section 4: Income by Treatment (helper table L:N) ----
    title_cell(ws_d, "L2", "Income by Treatment", size=13)
    for ci, h in enumerate(["Treatment", "This Month", "YTD"], start=12):
        c = ws_d.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=PINK)
    # Treatment names come from treatments.json (her treatments app), so the
    # per-treatment breakdown totals against exactly the treatments she's set up.
    # Any blank slots (if she has fewer than TREATMENT_SLOTS) total to 0 harmlessly.
    treatment_names = load_treatment_names()
    for si in range(TREATMENT_SLOTS):
        r = 4 + si
        name = treatment_names[si] if si < len(treatment_names) else ""
        ws_d.cell(row=r, column=12, value=name)
        ws_d.cell(row=r, column=13,
                  value=f'=IF($L{r}="",0,SUMPRODUCT(--({IN_TREAT}=$L{r}),--(LEFT({IN_DATE},7)=$B$4),N(+{IN_EUR})))'
                  ).number_format = "#,##0.00"
        ws_d.cell(row=r, column=14,
                  value=f'=IF($L{r}="",0,SUMPRODUCT(--({IN_TREAT}=$L{r}),--(LEFT({IN_DATE},4)=$B$14),N(+{IN_EUR})))'
                  ).number_format = "#,##0.00"
    label(ws_d, "L23", "Most popular (by count)")
    ws_d["M23"] = '=IF(COUNTA($L$4:$L$21)=0,"—",INDEX($L$4:$L$21,MATCH(MAX($N$4:$N$21),$N$4:$N$21,0)))'
    label(ws_d, "L24", "Highest earning (YTD)")
    ws_d["M24"] = '=IF(SUM($N$4:$N$21)=0,"—",INDEX($L$4:$L$21,MATCH(MAX($N$4:$N$21),$N$4:$N$21,0)))'

    # ---- Section 5: Business vs Personal by month (helper P:R) ----
    title_cell(ws_d, "A31", "Business vs Personal (Out, by month)", size=13)
    for ci, h in enumerate(["Month", "Business (EUR)", "Personal (EUR)"], start=1):
        c = ws_d.cell(row=32, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=PINK)
    for mi, mname in enumerate(MONTHS):
        r = 33 + mi
        mm = f"{mi+1:02d}"
        ws_d.cell(row=r, column=1, value=mname)
        ws_d.cell(row=r, column=2,
                  value=f'=SUMPRODUCT(--({OUT_BP}="Business"),--(LEFT({OUT_DATE},7)=($B$14&"-{mm}")),N(+{OUT_EUR}))'
                  ).number_format = "#,##0.00"
        ws_d.cell(row=r, column=3,
                  value=f'=SUMPRODUCT(--({OUT_BP}="Personal"),--(LEFT({OUT_DATE},7)=($B$14&"-{mm}")),N(+{OUT_EUR}))'
                  ).number_format = "#,##0.00"

    # ---- Section 6: Foreign currency spend grouped by currency ----
    title_cell(ws_d, "H26", "Foreign currency spend (Out)", size=13)
    for ci, h in enumerate(["Currency", "Orig Total", "EUR Total"], start=8):
        c = ws_d.cell(row=27, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor=PINK)
    common_curr = ["CHF", "USD", "GBP", "SEK", "NOK", "DKK", "PLN", "CAD", "AUD", "JPY"]
    for ci2, cur in enumerate(common_curr):
        r = 28 + ci2
        ws_d.cell(row=r, column=8, value=cur)
        ws_d.cell(row=r, column=9,
                  value=f'=SUMIF({OUT_CUR},$H{r},{OUT_TOTAL})').number_format = "#,##0.00"
        ws_d.cell(row=r, column=10,
                  value=f'=SUMIF({OUT_CUR},$H{r},{OUT_EUR})').number_format = "#,##0.00"

    # =================================================================
    # Charts — bind to the formula-driven helper tables above.
    # =================================================================
    # Chart 1: Monthly In vs Out (bar)
    c1 = BarChart(); c1.type = "col"; c1.title = "Monthly In vs Out (EUR)"; c1.style = 10
    data = Reference(ws_d, min_col=2, max_col=3, min_row=15, max_row=27)
    cats = Reference(ws_d, min_col=1, min_row=16, max_row=27)
    c1.add_data(data, titles_from_data=True); c1.set_categories(cats)
    c1.height = 7.5; c1.width = 15
    ws_d.add_chart(c1, "A45")

    # Chart 2: Cumulative net (line) — uses the Net column; cumulative is approximated
    # by plotting monthly net. (REVISIT: true running total needs a helper column.)
    c2 = LineChart(); c2.title = "Monthly Net (EUR)"; c2.style = 12
    data2 = Reference(ws_d, min_col=4, max_col=4, min_row=15, max_row=27)
    c2.add_data(data2, titles_from_data=True); c2.set_categories(cats)
    c2.height = 7.5; c2.width = 15
    ws_d.add_chart(c2, "J45")

    # Chart 3: Expense by category (donut) — YTD column
    c3 = DoughnutChart(); c3.title = "Expenses by Article (YTD)"
    data3 = Reference(ws_d, min_col=10, min_row=3, max_row=21)
    cats3 = Reference(ws_d, min_col=8, min_row=4, max_row=21)
    c3.add_data(data3, titles_from_data=True); c3.set_categories(cats3)
    c3.height = 8; c3.width = 11
    ws_d.add_chart(c3, "A62")

    # Chart 4: Income by treatment (donut) — YTD column
    c4 = DoughnutChart(); c4.title = "Income by Treatment (YTD)"
    data4 = Reference(ws_d, min_col=14, min_row=3, max_row=21)
    cats4 = Reference(ws_d, min_col=12, min_row=4, max_row=21)
    c4.add_data(data4, titles_from_data=True); c4.set_categories(cats4)
    c4.height = 8; c4.width = 11
    ws_d.add_chart(c4, "F62")

    # Chart 5: Business vs Personal by month (stacked bar)
    c5 = BarChart(); c5.type = "col"; c5.grouping = "stacked"; c5.overlap = 100
    c5.title = "Business vs Personal by month (EUR)"; c5.style = 11
    data5 = Reference(ws_d, min_col=2, max_col=3, min_row=32, max_row=44)
    cats5 = Reference(ws_d, min_col=1, min_row=33, max_row=44)
    c5.add_data(data5, titles_from_data=True); c5.set_categories(cats5)
    c5.height = 7.5; c5.width = 15
    ws_d.add_chart(c5, "J62")

    wb.save(EXCEL_PATH)
    print(f"Created {EXCEL_PATH}")


if __name__ == "__main__":
    build()
